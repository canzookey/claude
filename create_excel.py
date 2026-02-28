import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== Style definitions =====
title_font = Font(name="Yu Gothic", size=16, bold=True, color="FFFFFF")
h2_font = Font(name="Yu Gothic", size=13, bold=True, color="1F4E79")
h3_font = Font(name="Yu Gothic", size=11, bold=True, color="2E75B6")
header_font = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
body_font = Font(name="Yu Gothic", size=10)
bold_font = Font(name="Yu Gothic", size=10, bold=True)
number_font = Font(name="Yu Gothic", size=10)

title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
accent_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
warning_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

wrap_align = Alignment(wrap_text=True, vertical="center")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
right_align = Alignment(horizontal="right", vertical="center")


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_title(ws, row, title, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 36
    return row + 1


def write_section(ws, row, title, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = h2_font
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 28
    return row + 1


def write_subsection(ws, row, title, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = h3_font
    ws.row_dimensions[row].height = 24
    return row + 1


def write_table(ws, row, headers, data, col_start=1, bold_last_row=False, number_cols=None):
    if number_cols is None:
        number_cols = []
    # Header row
    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1
    # Data rows
    for i, data_row in enumerate(data):
        is_last = (i == len(data) - 1) and bold_last_row
        for j, val in enumerate(data_row):
            cell = ws.cell(row=row, column=col_start + j, value=val)
            cell.font = bold_font if is_last else body_font
            cell.border = thin_border
            cell.alignment = center_align if j > 0 else wrap_align
            if is_last:
                cell.fill = highlight_fill
            elif i % 2 == 0:
                cell.fill = light_fill
        row += 1
    return row


# ============================================================
# Sheet 1: エグゼクティブサマリー
# ============================================================
ws1 = wb.active
ws1.title = "1. エグゼクティブサマリー"
set_col_widths(ws1, [30, 20, 20, 20, 20])

r = 1
r = write_title(ws1, r, "SaaS 事業計画書")
r += 1

r = write_section(ws1, r, "1. エグゼクティブサマリー")
r += 1

r = write_subsection(ws1, r, "1.1 事業概要")
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws1.cell(row=r, column=1, value="中小企業向けクラウドベース業務効率化SaaSプラットフォーム。プロジェクト管理、CRM、請求・経理機能を統合提供し、中小企業のDX推進を支援する。").font = body_font
ws1.cell(row=r, column=1).alignment = wrap_align
ws1.row_dimensions[r].height = 40
r += 2

r = write_subsection(ws1, r, "1.2 ミッション")
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws1.cell(row=r, column=1, value="「すべての中小企業にシンプルで強力なデジタルツールを届ける」").font = Font(name="Yu Gothic", size=11, bold=True, italic=True)
r += 2

r = write_subsection(ws1, r, "1.3 ビジョン")
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws1.cell(row=r, column=1, value="2030年までに国内中小企業向けSaaS市場においてトップ3のポジションを確立する。").font = body_font
r += 2

r = write_subsection(ws1, r, "1.4 主要KPI目標（3年計画）")
r = write_table(ws1, r,
    ["指標", "1年目", "2年目", "3年目"],
    [
        ["ARR（年間経常収益）", "1.2億円", "4.8億円", "12億円"],
        ["有料顧客数", "200社", "700社", "1,500社"],
        ["月次解約率（Churn Rate）", "< 3.0%", "< 2.0%", "< 1.5%"],
        ["NRR（売上継続率）", "105%", "115%", "125%"],
    ])

# ============================================================
# Sheet 2: 市場分析
# ============================================================
ws2 = wb.create_sheet("2. 市場分析")
set_col_widths(ws2, [25, 25, 25, 25])

r = 1
r = write_title(ws2, r, "2. 市場分析", 4)
r += 1

r = write_subsection(ws2, r, "2.1 市場規模", 4)
data = [
    ["国内SaaS市場規模", "約1.8兆円（2025年）、年率13%成長"],
    ["中小企業向けセグメント", "約4,500億円"],
    ["ターゲット市場（SAM）", "約800億円（従業員10〜300名規模）"],
]
for item in data:
    ws2.cell(row=r, column=1, value=item[0]).font = bold_font
    ws2.cell(row=r, column=1).border = thin_border
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws2.cell(row=r, column=2, value=item[1]).font = body_font
    ws2.cell(row=r, column=2).border = thin_border
    r += 1
r += 1

r = write_subsection(ws2, r, "2.2 市場トレンド", 4)
trends = [
    "DX推進の加速: 政府のデジタル化推進政策により中小企業のIT投資が増加",
    "クラウドシフト: オンプレミスからクラウドへの移行が加速",
    "統合プラットフォーム需要: 複数ツールの乱立による非効率を解消したいニーズ",
    "AI活用の普及: 業務自動化・データ分析へのAI統合が差別化要因に",
]
for i, t in enumerate(trends, 1):
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws2.cell(row=r, column=1, value=f"{i}. {t}").font = body_font
    r += 1
r += 1

r = write_subsection(ws2, r, "2.3 競合分析", 4)
r = write_table(ws2, r,
    ["競合", "強み", "弱み", "ポジション"],
    [
        ["A社（大手SaaS）", "ブランド力、機能豊富", "高価格、カスタマイズ性低い", "エンタープライズ向け"],
        ["B社（CRM特化）", "CRM機能の深さ", "業務統合が限定的", "CRM特化"],
        ["C社（会計SaaS）", "会計機能の完成度", "プロジェクト管理機能なし", "会計特化"],
        ["★ 当社", "統合性、価格競争力、AI活用", "ブランド認知度", "中小企業統合型"],
    ])
# Highlight our row
for col in range(1, 5):
    ws2.cell(row=r-1, column=col).fill = accent_fill
    ws2.cell(row=r-1, column=col).font = bold_font
r += 1

r = write_subsection(ws2, r, "2.4 ターゲット顧客", 4)
targets = [
    ["主要ターゲット", "従業員10〜100名の中小企業"],
    ["業種", "IT/Web、コンサルティング、製造業、サービス業"],
    ["課題", "複数ツールの管理負荷、データの分断、高額なIT投資"],
]
for item in targets:
    ws2.cell(row=r, column=1, value=item[0]).font = bold_font
    ws2.cell(row=r, column=1).border = thin_border
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws2.cell(row=r, column=2, value=item[1]).font = body_font
    ws2.cell(row=r, column=2).border = thin_border
    r += 1

# ============================================================
# Sheet 3: プロダクト戦略
# ============================================================
ws3 = wb.create_sheet("3. プロダクト戦略")
set_col_widths(ws3, [20, 40, 25, 25])

r = 1
r = write_title(ws3, r, "3. プロダクト戦略", 4)
r += 1

r = write_subsection(ws3, r, "3.1 コア機能ロードマップ", 4)
r = write_table(ws3, r,
    ["フェーズ", "期間", "主要機能", "目的"],
    [
        ["フェーズ1 (MVP)", "0〜6ヶ月", "プロジェクト管理、基本CRM、ダッシュボード", "市場投入・検証"],
        ["フェーズ2 (成長期)", "7〜12ヶ月", "請求書発行、ワークフロー自動化、API連携", "機能拡充・定着"],
        ["フェーズ3 (拡大期)", "13〜24ヶ月", "AI分析・予測、BIレポート、モバイルアプリ", "差別化・拡大"],
        ["フェーズ4 (成熟期)", "25〜36ヶ月", "マーケットプレイス、業種別テンプレート、エンタープライズ機能", "エコシステム構築"],
    ])
r += 1

r = write_subsection(ws3, r, "3.2 技術スタック", 4)
r = write_table(ws3, r,
    ["レイヤー", "技術", "", ""],
    [
        ["フロントエンド", "React / Next.js / TypeScript", "", ""],
        ["バックエンド", "Node.js / NestJS", "", ""],
        ["データベース", "PostgreSQL / Redis", "", ""],
        ["インフラ", "AWS (ECS, RDS, S3, CloudFront)", "", ""],
        ["CI/CD", "GitHub Actions", "", ""],
        ["モニタリング", "Datadog / Sentry", "", ""],
        ["AI/ML", "Python / OpenAI API / 自社モデル", "", ""],
    ])
r += 1

r = write_subsection(ws3, r, "3.3 差別化ポイント", 4)
diffs = [
    ["オールインワン統合", "1つのプラットフォームで主要業務をカバー"],
    ["AI搭載", "業務データからのインサイト自動生成"],
    ["価格競争力", "個別ツール3つ分の半額以下"],
    ["日本市場最適化", "日本の商習慣（請求書、見積書、稟議）に完全対応"],
]
for item in diffs:
    ws3.cell(row=r, column=1, value=item[0]).font = bold_font
    ws3.cell(row=r, column=1).border = thin_border
    ws3.cell(row=r, column=1).fill = accent_fill
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws3.cell(row=r, column=2, value=item[1]).font = body_font
    ws3.cell(row=r, column=2).border = thin_border
    r += 1

# ============================================================
# Sheet 4: ビジネスモデル
# ============================================================
ws4 = wb.create_sheet("4. ビジネスモデル")
set_col_widths(ws4, [20, 22, 18, 40])

r = 1
r = write_title(ws4, r, "4. ビジネスモデル", 4)
r += 1

r = write_subsection(ws4, r, "4.1 料金プラン", 4)
r = write_table(ws4, r,
    ["プラン", "月額（1ユーザー）", "対象", "主な機能"],
    [
        ["Free", "¥0", "〜3名", "基本機能制限版（リード獲得用）"],
        ["Starter", "¥1,980", "〜10名", "基本機能、5GBストレージ"],
        ["Business", "¥3,980", "〜50名", "全機能、50GBストレージ、API連携"],
        ["Enterprise", "¥7,980", "50名〜", "全機能、無制限ストレージ、SSO、専任サポート"],
    ])
r += 1

r = write_subsection(ws4, r, "4.2 収益モデル", 4)
rev_models = [
    ["サブスクリプション収益", "月額/年額課金（年額は10%割引）"],
    ["アドオン収益", "AI機能パック、追加ストレージ、カスタムレポート"],
    ["プロフェッショナルサービス", "導入支援、データ移行、カスタム開発"],
]
for item in rev_models:
    ws4.cell(row=r, column=1, value=item[0]).font = bold_font
    ws4.cell(row=r, column=1).border = thin_border
    ws4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws4.cell(row=r, column=2, value=item[1]).font = body_font
    ws4.cell(row=r, column=2).border = thin_border
    r += 1
r += 1

r = write_subsection(ws4, r, "4.3 ユニットエコノミクス（目標値）", 4)
r = write_table(ws4, r,
    ["指標", "1年目", "2年目", "3年目"],
    [
        ["ARPU（月額客単価）", "¥50,000", "¥57,000", "¥67,000"],
        ["CAC（顧客獲得コスト）", "¥300,000", "¥250,000", "¥200,000"],
        ["LTV（顧客生涯価値）", "¥1,200,000", "¥1,710,000", "¥2,680,000"],
        ["LTV/CAC比率", "4.0x", "6.8x", "13.4x"],
        ["CAC回収期間", "6ヶ月", "4.4ヶ月", "3.0ヶ月"],
    ])

# ============================================================
# Sheet 5: マーケティング・営業戦略
# ============================================================
ws5 = wb.create_sheet("5. マーケティング・営業")
set_col_widths(ws5, [22, 20, 20, 20, 20])

r = 1
r = write_title(ws5, r, "5. マーケティング・営業戦略")
r += 1

r = write_subsection(ws5, r, "5.1 Go-to-Market戦略")
phases = [
    ["フェーズ1: PLG", "フリープランによるオーガニック獲得、コンテンツマーケティング（SEO）、SNS（X, LinkedIn, YouTube）"],
    ["フェーズ2: セールス強化", "インサイドセールスチーム構築、パートナーチャネル（SIer、税理士事務所）、ウェビナー"],
    ["フェーズ3: エンタープライズ", "フィールドセールス、大型案件対応、戦略的アライアンス"],
]
for item in phases:
    ws5.cell(row=r, column=1, value=item[0]).font = bold_font
    ws5.cell(row=r, column=1).border = thin_border
    ws5.cell(row=r, column=1).fill = light_fill
    ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws5.cell(row=r, column=2, value=item[1]).font = body_font
    ws5.cell(row=r, column=2).border = thin_border
    ws5.cell(row=r, column=2).alignment = wrap_align
    ws5.row_dimensions[r].height = 35
    r += 1
r += 1

r = write_subsection(ws5, r, "5.2 マーケティング予算配分")
r = write_table(ws5, r,
    ["チャネル", "1年目", "2年目", "3年目"],
    [
        ["デジタル広告", "40%", "30%", "20%"],
        ["コンテンツ/SEO", "25%", "25%", "20%"],
        ["イベント/ウェビナー", "15%", "20%", "20%"],
        ["パートナー", "10%", "15%", "25%"],
        ["PR/ブランディング", "10%", "10%", "15%"],
    ])
r += 1

r = write_subsection(ws5, r, "5.3 カスタマーサクセス")
cs_items = [
    ["オンボーディング", "専用CSMによる導入支援（Businessプラン以上）"],
    ["ヘルススコア管理", "利用状況に基づく離脱予防"],
    ["アップセル/クロスセル", "利用状況に応じた上位プラン・アドオン提案"],
    ["コミュニティ", "ユーザーコミュニティ運営、勉強会開催"],
]
for item in cs_items:
    ws5.cell(row=r, column=1, value=item[0]).font = bold_font
    ws5.cell(row=r, column=1).border = thin_border
    ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws5.cell(row=r, column=2, value=item[1]).font = body_font
    ws5.cell(row=r, column=2).border = thin_border
    r += 1

# ============================================================
# Sheet 6: 組織計画
# ============================================================
ws6 = wb.create_sheet("6. 組織計画")
set_col_widths(ws6, [30, 18, 18, 18])

r = 1
r = write_title(ws6, r, "6. 組織計画", 4)
r += 1

r = write_subsection(ws6, r, "6.1 チーム構成", 4)
r = write_table(ws6, r,
    ["部門", "1年目", "2年目", "3年目"],
    [
        ["経営", "2名", "3名", "4名"],
        ["エンジニアリング", "8名", "15名", "25名"],
        ["プロダクト/デザイン", "2名", "4名", "6名"],
        ["セールス", "3名", "8名", "15名"],
        ["カスタマーサクセス", "2名", "5名", "10名"],
        ["マーケティング", "2名", "4名", "7名"],
        ["管理（人事/経理/法務）", "1名", "3名", "5名"],
        ["合計", "20名", "42名", "72名"],
    ], bold_last_row=True)
r += 1

r = write_subsection(ws6, r, "6.2 採用戦略", 4)
strategies = [
    "テックカンファレンスでの登壇・スポンサー",
    "OSS活動によるエンジニアブランディング",
    "リファラル採用プログラム（紹介報酬制度）",
    "ストックオプション制度による人材確保",
]
for s in strategies:
    ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws6.cell(row=r, column=1, value=f"• {s}").font = body_font
    r += 1

# ============================================================
# Sheet 7: 財務計画
# ============================================================
ws7 = wb.create_sheet("7. 財務計画")
set_col_widths(ws7, [30, 18, 18, 18])

r = 1
r = write_title(ws7, r, "7. 財務計画", 4)
r += 1

r = write_subsection(ws7, r, "7.1 売上計画", 4)
r = write_table(ws7, r,
    ["項目", "1年目", "2年目", "3年目"],
    [
        ["サブスクリプション収益", "1.0億円", "4.2億円", "10.5億円"],
        ["アドオン収益", "0.1億円", "0.3億円", "0.9億円"],
        ["プロフェッショナルサービス", "0.1億円", "0.3億円", "0.6億円"],
        ["売上合計", "1.2億円", "4.8億円", "12.0億円"],
    ], bold_last_row=True)
r += 1

r = write_subsection(ws7, r, "7.2 費用計画", 4)
r = write_table(ws7, r,
    ["項目", "1年目", "2年目", "3年目"],
    [
        ["人件費", "2.0億円", "4.2億円", "7.2億円"],
        ["インフラ/ホスティング", "0.3億円", "0.6億円", "1.2億円"],
        ["マーケティング", "0.8億円", "1.5億円", "2.4億円"],
        ["オフィス/管理費", "0.3億円", "0.5億円", "0.8億円"],
        ["その他（法務、ツール等）", "0.2億円", "0.3億円", "0.4億円"],
        ["費用合計", "3.6億円", "7.1億円", "12.0億円"],
    ], bold_last_row=True)
r += 1

r = write_subsection(ws7, r, "7.3 損益計画", 4)
r = write_table(ws7, r,
    ["項目", "1年目", "2年目", "3年目"],
    [
        ["売上", "1.2億円", "4.8億円", "12.0億円"],
        ["売上原価（インフラ等）", "0.3億円", "0.6億円", "1.2億円"],
        ["粗利益", "0.9億円", "4.2億円", "10.8億円"],
        ["粗利率", "75%", "87.5%", "90%"],
        ["営業費用", "3.3億円", "6.5億円", "10.8億円"],
        ["営業利益", "-2.4億円", "-2.3億円", "0億円"],
        ["営業利益率", "-200%", "-48%", "0%"],
    ])
# Highlight negative rows
for offset in [5, 6]:  # 営業利益 and 営業利益率 rows
    target_row = r - 2 + offset - 5
r += 1

r = write_subsection(ws7, r, "7.4 資金調達計画", 4)
r = write_table(ws7, r,
    ["ラウンド", "時期", "調達額", "用途"],
    [
        ["シード", "創業時", "1.0億円", "MVP開発、初期チーム"],
        ["シリーズA", "1年目末", "5.0億円", "プロダクト強化、セールス拡大"],
        ["シリーズB", "3年目", "15.0億円", "市場拡大、エンタープライズ展開"],
    ])
r += 1

r = write_subsection(ws7, r, "7.5 キャッシュフロー", 4)
r = write_table(ws7, r,
    ["項目", "1年目", "2年目", "3年目"],
    [
        ["期首キャッシュ", "1.0億円", "3.6億円", "6.3億円"],
        ["営業CF", "-2.4億円", "-2.3億円", "0億円"],
        ["資金調達", "5.0億円", "5.0億円", "15.0億円"],
        ["期末キャッシュ", "3.6億円", "6.3億円", "21.3億円"],
    ], bold_last_row=True)

# ============================================================
# Sheet 8: リスクと対策
# ============================================================
ws8 = wb.create_sheet("8. リスクと対策")
set_col_widths(ws8, [28, 12, 12, 48])

r = 1
r = write_title(ws8, r, "8. リスクと対策", 4)
r += 1

r = write_subsection(ws8, r, "8.1 主要リスク", 4)
r = write_table(ws8, r,
    ["リスク", "影響度", "発生確率", "対策"],
    [
        ["大手SaaSの中小企業市場参入", "高", "中", "日本市場特化、迅速な機能開発、価格競争力維持"],
        ["顧客獲得コストの高騰", "中", "高", "PLG戦略強化、パートナーチャネル多様化"],
        ["エンジニア採用難", "高", "高", "技術ブランディング、リモートワーク対応、SO制度"],
        ["セキュリティインシデント", "高", "低", "SOC2取得、定期ペネトレーションテスト、バグバウンティ"],
        ["景気後退によるIT投資縮小", "中", "中", "コスト削減効果の訴求、柔軟な料金プラン"],
    ])
r += 1

r = write_subsection(ws8, r, "8.2 規制・コンプライアンス対応", 4)
compliance = [
    "個人情報保護法対応",
    "ISMAP（政府情報システムのセキュリティ評価制度）取得検討",
    "SOC2 Type II 認証取得（2年目）",
    "データセンター: 国内リージョン利用",
]
for c in compliance:
    ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws8.cell(row=r, column=1, value=f"• {c}").font = body_font
    r += 1

# ============================================================
# Sheet 9: マイルストーン
# ============================================================
ws9 = wb.create_sheet("9. マイルストーン")
set_col_widths(ws9, [20, 50, 20])

r = 1
r = write_title(ws9, r, "9. マイルストーン", 3)
r += 1

r = write_table(ws9, r,
    ["時期", "マイルストーン", "年度"],
    [
        ["Q1 Y1", "MVP完成、クローズドベータ開始", "1年目"],
        ["Q2 Y1", "パブリックベータ、有料プラン開始", "1年目"],
        ["Q3 Y1", "有料顧客100社達成", "1年目"],
        ["Q4 Y1", "シリーズA調達完了、ARR 1億円突破", "1年目"],
        ["Q2 Y2", "フェーズ2機能リリース、API連携開始", "2年目"],
        ["Q4 Y2", "有料顧客500社達成、ARR 4億円突破", "2年目"],
        ["Q2 Y3", "AI機能リリース、モバイルアプリリリース", "3年目"],
        ["Q4 Y3", "有料顧客1,500社、ARR 12億円、損益分岐点到達", "3年目"],
    ])

# ============================================================
# Sheet 10: 用語集・前提条件
# ============================================================
ws10 = wb.create_sheet("10. 付録")
set_col_widths(ws10, [20, 50])

r = 1
r = write_title(ws10, r, "10. 付録", 2)
r += 1

r = write_subsection(ws10, r, "10.1 用語集", 2)
r = write_table(ws10, r,
    ["用語", "説明"],
    [
        ["ARR", "Annual Recurring Revenue（年間経常収益）"],
        ["MRR", "Monthly Recurring Revenue（月間経常収益）"],
        ["ARPU", "Average Revenue Per User（平均顧客単価）"],
        ["CAC", "Customer Acquisition Cost（顧客獲得コスト）"],
        ["LTV", "Lifetime Value（顧客生涯価値）"],
        ["NRR", "Net Revenue Retention（売上継続率）"],
        ["Churn Rate", "解約率"],
        ["PLG", "Product-Led Growth（プロダクト主導型成長）"],
        ["CSM", "Customer Success Manager（カスタマーサクセスマネージャー）"],
        ["SO", "Stock Option（ストックオプション）"],
    ])
r += 1

r = write_subsection(ws10, r, "10.2 前提条件", 2)
assumptions = [
    ["月次顧客成長率", "15%（1年目）、12%（2年目）、8%（3年目）"],
    ["平均契約ユーザー数", "10名/社"],
    ["年間契約比率", "60%"],
    ["粗利率", "75%〜90%（スケールに伴い改善）"],
    ["為替レート", "考慮なし（国内市場前提）"],
]
for item in assumptions:
    ws10.cell(row=r, column=1, value=item[0]).font = bold_font
    ws10.cell(row=r, column=1).border = thin_border
    ws10.cell(row=r, column=2, value=item[1]).font = body_font
    ws10.cell(row=r, column=2).border = thin_border
    r += 1
r += 2
ws10.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws10.cell(row=r, column=1, value="本事業計画書は2026年2月時点の情報に基づいて作成されています。").font = Font(name="Yu Gothic", size=9, italic=True, color="808080")

# Save
wb.save("/home/user/claude/saas-business-plan.xlsx")
print("Excel file created: saas-business-plan.xlsx")
